import base64
import pandas as pd
from io import BytesIO
from openpyxl import load_workbook
from odoo import api, fields, models, _
from odoo.exceptions import UserError, ValidationError
from datetime import date, datetime
from datetime import datetime, timedelta
import logging

_logger = logging.getLogger(__name__)


class PayslipExcelImport(models.Model):
    _name = 'payslip.excel.import'
    _description = 'Excel Payslip'
    _inherit = ['mail.thread.cc', 'mail.thread.main.attachment', 'mail.activity.mixin']

    name = fields.Char(string='Ref')
    user_id = fields.Many2one('res.users', string='Requested By', default=lambda self: self.env.user, readonly=True)
    company_id = fields.Many2one('res.company', default=lambda self: self.env.company)
    currency_symbol = fields.Char(
        string='Currency Symbol', default=lambda self: self.env.company.currency_id.symbol, readonly=True, )
    logged_user = fields.Many2one('res.users', string='Logged User', compute='get_logged_user')
    active = fields.Boolean(default=True)
    status = fields.Selection([('draft', 'Draft'), ('generate', 'Payslip Generated')],
                              string="State",
                              default='draft')
    is_executed = fields.Boolean(string='Is Executed')
    excel_sheet = fields.Binary(string='Upload Excel Sheet')
    data_html = fields.Html(string='Excel Data', readonly=True, help="HTML representation of the Excel data")

    select_month = fields.Selection([
        ('January', 'January'),
        ('February', 'February'),
        ('March', 'March'),
        ('April', 'April'),
        ('May', 'May'),
        ('June', 'June'),
        ('July', 'July'),
        ('August', 'August'),
        ('September', 'September'),
        ('October', 'October'),
        ('November', 'November'),
        ('December', 'December'),
    ], string="Month", default=lambda self: datetime.now().strftime('%B'))
    select_year = fields.Many2one('hr.payroll.year', string='Year ',
                                  default=lambda self: self._default_year())
    pay_date = fields.Date(string='Pay Date', default=fields.Date.today())
    date_from = fields.Date(
        string='From', compute='_compute_dates')
    date_to = fields.Date(
        string='To', compute='_compute_dates')
    payslip_ids = fields.Many2many('hr.payslip', string='Payslip Reference')
    payslip_count = fields.Integer(string='Payslip Count', compute='_compute_payslip_count')

    total_days_of_month = fields.Html(string=' ', compute='_compute_days_and_leaves')
    attendance_type = fields.Selection([
        ('automatic', 'Automatic'),
        ('manual', 'Manual'), ], string='Type')

    @api.onchange('select_year', 'select_month')
    def _compute_days_and_leaves(self):
        for rec in self:
            if rec.select_year and rec.select_month:
                total_days = 0.00
                worker = 0.00
                staff_pay = 0.00
                staff_pay_sat = 0.00
                saturday = 0.00
                sunday = 0.00
                public_holidays = 0.00
                month_names = {
                    1: 'January', 2: 'February', 3: 'March', 4: 'April',
                    5: 'May', 6: 'June', 7: 'July', 8: 'August',
                    9: 'September', 10: 'October', 11: 'November', 12: 'December'}
                for year in rec.select_year.day_and_month:
                    if month_names.get(int(year.select_month)) == rec.select_month:
                        total_days += year.total_number_of_days
                        worker += year.number_of_days + year.only_saturday - year.public_holiday_count
                        staff_pay += year.number_of_days + year.staff_saturday - year.public_holiday_count
                        staff_pay_sat += year.number_of_days + year.only_saturday - year.public_holiday_count
                        saturday += year.only_saturday
                        sunday += year.only_sunday
                        public_holidays += year.public_holiday_count

                html_content = f"""
                <div style="padding: 10px;padding-bottom: 0px; text-align: center;">
                    <div style="padding: 10px; border: 1px solid #ddd; border-radius: 0px; background-color: #f9f9f9;">
                        <h4 style="color: #212529; margin: 0;">Total Days in Month: <span style="font-weight: normal;">{total_days} Days</span></h4>
                    </div>
                </div>
                <div style="display: flex; padding: 0 10px;">
                    <div style="flex: 1; padding: 16px; border: 1px solid #ddd; border-radius: 0px; background-color: #f9f9f9;">
                        <h4 style="color: #212529; margin-bottom: 10px;">Weekoff & Public Holidays</h4>
                        <hr style="border: 0; border-top: 2px solid #ddd; margin: 10px 0;" />
                        <p style="font-size: 16px; font-weight: 500; color: #212529; margin: 5px 0;"><span>Saturday : </span> {saturday} Days</p>
                        <p style="font-size: 16px; font-weight: 500; color: #212529; margin: 5px 0;"><span>Sunday : </span> {sunday} Days</p>
                        <p style="font-size: 16px; font-weight: 500; color: #212529; margin: 5px 0;"><span>Public Holiday : </span> {public_holidays} Days</p>
                    </div>
                    <div style="flex: 1; padding: 16px; border: 1px solid #ddd; border-radius: 0px; background-color: #f9f9f9;">
                        <h4 style="color: #212529; margin-bottom: 10px;">Working Days</h4>
                        <hr style="border: 0; border-top: 2px solid #ddd; margin: 10px 0;" />
                        <p style="font-size: 16px; font-weight: 500; color: #212529; margin: 5px 0;"><span>Worker Pay : </span> {worker} Days</p>
                        <p style="font-size: 16px; font-weight: 500; color: #212529; margin: 5px 0;"><span>Staff Pay : </span> {staff_pay} Days</p>
                        <p style="font-size: 16px; font-weight: 500; color: #212529; margin: 5px 0;"><span>Staff Pay Sat : </span> {staff_pay_sat} Days</p>
                    </div>
                </div>
                """

                rec.total_days_of_month = html_content
            else:
                rec.total_days_of_month = '<p style="color: red;">Please select both year and month.</p>'

    def _compute_payslip_count(self):
        for record in self:
            record.payslip_count = self.env['hr.payslip'].search_count([('id', 'in', record.payslip_ids.ids)])

    def get_payslip_views(self):
        self.sudo().ensure_one()
        form_view = self.sudo().env.ref('hr_payroll.view_hr_payslip_form')
        tree_view = self.sudo().env.ref('hr_payroll.view_hr_payslip_tree')
        return {
            'name': _('Payslip'),
            'res_model': 'hr.payslip',
            'type': 'ir.actions.act_window',
            'view_mode': 'tree,form',
            'views': [(tree_view.id, 'tree'), (form_view.id, 'form')],
            'domain': [('id', 'in', self.payslip_ids.ids)],
        }

    @api.depends('select_month', 'select_year')
    def _compute_dates(self):
        for record in self:
            if record.select_month:
                current_year = datetime.now().year
                year = int(record.select_year.name) if record.select_year and record.select_year.name else current_year
                month = datetime.strptime(record.select_month, '%B').month
                date_to = date(year, month, 25)
                if month == 1:
                    prev_month_year = year - 1
                    prev_month = 12
                else:
                    prev_month_year = year
                    prev_month = month - 1
                date_from = date(prev_month_year, prev_month, 26)
                record.date_from = date_from
                record.date_to = date_to

    @api.model
    def _default_year(self):
        current_year = datetime.now().year
        year = self.env['hr.payroll.year'].search([('name', '=', str(current_year))])
        return year and year.id or False

    def get_logged_user(self):
        self.logged_user = self.env.uid

    def get_excel_data(self):
        self.ensure_one()
        if not self.excel_sheet:
            raise UserError(_("Please upload an Excel sheet first."))
        else:
            excel_data = BytesIO(base64.b64decode(self.excel_sheet))
            try:
                wb = load_workbook(excel_data)
                for ws in wb.worksheets:
                    for row in ws.iter_rows():
                        for cell in row:
                            cell.style = 'Normal'
                            if isinstance(cell.value, str) and cell.value.endswith('px'):
                                cell.value = cell.value.replace('px', '')

                cleaned_data = BytesIO()
                wb.save(cleaned_data)
                cleaned_data.seek(0)
                df = pd.read_excel(cleaned_data)
            except Exception as e:
                raise UserError(_("Error reading Excel file: %s") % str(e))

            df.fillna(0, inplace=True)

            def excel_date_to_str(excel_date):
                if pd.isna(excel_date) or excel_date == 0:
                    return ''
                try:
                    return (datetime(1899, 12, 30) + timedelta(days=excel_date)).strftime('%Y-%m-%d')
                except Exception as e:
                    print(f"Error converting date: {excel_date} - {str(e)}")
                    return ''

            if 'Leave/Start Date' in df.columns:
                df['Leave/Start Date'] = df['Leave/Start Date'].apply(excel_date_to_str)

            if 'Leave/End Date' in df.columns:
                df['Leave/End Date'] = df['Leave/End Date'].apply(excel_date_to_str)

            data_dict = df.to_dict(orient='records')
            table_header = ''
            table_datas = ''
            keys = df.columns.tolist()

            for i, key in enumerate(keys):
                if i == 0:
                    table_header += "<th style='border:1px solid black !important; position: sticky; left: 0; background-color: lightblue;'>{0}</th>".format(
                        key)
                elif i == 1:
                    table_header += "<th style='border:1px solid black !important; position: sticky; left: 130px; background-color: lightblue;'>{0}</th>".format(
                        key)
                else:
                    table_header += "<th style='border:1px solid black !important'>{0}</th>".format(key)
            for row in data_dict:
                table_datas += "<tr>"
                for i, key in enumerate(keys):
                    if i == 0:
                        table_datas += "<td style='border:1px solid black !important; position: sticky; left: 0; background-color: white;'>{0}</td>".format(
                            row[key])
                    elif i == 1:
                        table_datas += "<td style='border:1px solid black !important; position: sticky; left: 130px; background-color: white;'>{0}</td>".format(
                            row[key])
                    else:
                        table_datas += "<td style='border:1px solid black !important'>{0}</td>".format(row[key])
                table_datas += "</tr>"
            html_table = """
                <div style="overflow:auto;">
                    <table class="table text-center table-border table-sm" style="width:max-content;">
                        <thead>
                            <tr style='border:1px solid black !important;background: lightblue;'>
                                {header}
                            </tr>
                        </thead>
                        <tbody>
                            {data}
                        </tbody>
                    </table>
                </div>
            """.format(header=table_header, data=table_datas)
            self.data_html = html_table
            self.is_executed = True
            return data_dict

    #AUTOMATIC ATTENDANCE
    def generate_payslip_automatic_att(self):
        val = self.get_excel_data()
        new_dict = []
        current_employee = None
        for row in val:
            if row['Employee Code'] != 0:
                if current_employee:
                    new_dict.append(current_employee)
                current_employee = row.copy()
                current_employee['Leave'] = []
                current_employee.pop('Leave/Start Date', None)
                current_employee.pop('Leave/End Date', None)
                current_employee.pop('Leave/Half Day', None)
                current_employee.pop('Leave/Time Off Type', None)
                if row.get('Leave/Start Date'):
                    leave_row = {
                        'Leave/Start Date': row['Leave/Start Date'],
                        'Leave/End Date': row['Leave/End Date'],
                        'Leave/Half Day': row['Leave/Half Day'],
                        'Leave/Time Off Type': row['Leave/Time Off Type']
                    }
                    current_employee['Leave'].append(leave_row)
            else:
                leave_row = {
                    'Leave/Start Date': row['Leave/Start Date'],
                    'Leave/End Date': row['Leave/End Date'],
                    'Leave/Half Day': row['Leave/Half Day'],
                    'Leave/Time Off Type': row['Leave/Time Off Type']
                }
                if current_employee:
                    current_employee['Leave'].append(leave_row)
        if current_employee:
            new_dict.append(current_employee)
        for emp_data in new_dict:
            emp_code = emp_data['Employee Code']
            employees = self.env['hr.employee'].search([('emp_code', '=', emp_code)])
            if emp_code > 0 and not employees:
                raise UserError(_("Employee code %s not found in the system." % (
                    emp_code)))
            for employee in employees:
                payslip_name = self.env['ir.sequence'].next_by_code('salary.slip') or '/'
                time_off_data = []
                for leave in emp_data['Leave']:
                    leave_type = self.env['hr.leave.type'].search([('name', '=', leave['Leave/Time Off Type'])],
                                                                  limit=1)
                    if not leave_type:
                        raise UserError(_("Leave Type not found for Employee Code %s." % (
                            employee.emp_code)))
                    time_off_data.append((0, 0, {
                        'request_date_from': leave['Leave/Start Date'],
                        'request_date_to': leave['Leave/End Date'],
                        'request_unit_half': True if leave['Leave/Half Day'] == 0.5 else False,
                        'employee_id': employee.id,
                        'holiday_status_id': leave_type.id,
                    }))
                attendance_count = self.sudo().env['hr.attendance'].search([
                    ('employee_id', '=', employee.id)])
                count_val = 0
                for at in attendance_count:
                    if self.date_from <= at.check_in.date() <= self.date_to:
                        count_val += 1
                payslip = self.env['hr.payslip'].create({
                    'name': payslip_name,
                    'select_month': self.select_month,
                    'select_year': self.select_year.id,
                    'date_from': self.date_from,
                    'date_to': self.date_to,
                    'employee_id': employee.id,
                    'emp_code': employee.emp_code,
                    'attendance_type': 'automatic',
                    'employee_manual_present_days': count_val,
                    'extra_worked_days': emp_data['Extra Worked Days'],
                    'other_allowance': emp_data['Other Allowance'],
                    'attendance_bonus': emp_data['Attendance Bonus'],
                    'food_allowance': emp_data['Food Allowance'],
                    'incentive': emp_data['Incentive'],
                    'leave_encashment': emp_data['Leave Encashment'],
                    'overtime_hours': emp_data['Overtime Hours'],
                    'night_shift_allowance': emp_data['Night shift Allowance per day'],
                    'total_days_worked': emp_data['Total Night shift days worked'],
                    'production_incentive': emp_data['Production Incentive'],
                    'other_loan': emp_data['Other Loan'],
                    'food_deduction': emp_data['Food Deduction'],
                    'other_deduction': emp_data['Other Deduction'],
                    'room_rent_deduction': emp_data['Room Rent Deduction'],
                    'shoe_and_uniform_deduction': emp_data['Shoe & Uniform Deduction'],
                    'income_tax': emp_data['Income Tax'],
                    'actual_professional_tax': emp_data['Actual Professional Tax'],
                    'tds': emp_data['TDS'],
                    'time_off_ids': time_off_data,
                })
                payslip.employee_days_onchange()
                payslip.compute_sheet()
                payslip.compute_sheet()
                payslip.compute_sheet()
                payslip.compute_sheet()
                payslip.compute_sheet()
                self.write({
                    'payslip_ids': [(4, payslip.id)],
                    'status': 'generate'
                })


    #MANUAL ATTENDANCE
    def generate_payslip(self):
        val = self.get_excel_data()
        new_dict = []
        current_employee = None
        for row in val:
            if row['Employee Code'] != 0:
                if current_employee:
                    new_dict.append(current_employee)
                current_employee = row.copy()
                current_employee['Leave'] = []
                current_employee.pop('Leave/Start Date', None)
                current_employee.pop('Leave/End Date', None)
                current_employee.pop('Leave/Half Day', None)
                current_employee.pop('Leave/Time Off Type', None)
                if row.get('Leave/Start Date'):
                    leave_row = {
                        'Leave/Start Date': row['Leave/Start Date'],
                        'Leave/End Date': row['Leave/End Date'],
                        'Leave/Half Day': row['Leave/Half Day'],
                        'Leave/Time Off Type': row['Leave/Time Off Type']
                    }
                    current_employee['Leave'].append(leave_row)
            else:
                leave_row = {
                    'Leave/Start Date': row['Leave/Start Date'],
                    'Leave/End Date': row['Leave/End Date'],
                    'Leave/Half Day': row['Leave/Half Day'],
                    'Leave/Time Off Type': row['Leave/Time Off Type']
                }
                if current_employee:
                    current_employee['Leave'].append(leave_row)
        if current_employee:
            new_dict.append(current_employee)
        for emp_data in new_dict:
            emp_code = emp_data['Employee Code']
            employees = self.env['hr.employee'].search([('emp_code', '=', emp_code)])
            if emp_code > 0 and not employees:
                raise UserError(_("Employee code %s not found in the system." % (
                    emp_code)))
            for employee in employees:
                payslip_name = self.env['ir.sequence'].next_by_code('salary.slip') or '/'
                time_off_data = []
                for leave in emp_data['Leave']:
                    leave_type = self.env['hr.leave.type'].search([('name', '=', leave['Leave/Time Off Type'])],
                                                                  limit=1)
                    if not leave_type:
                        raise UserError(_("Leave Type not found for Employee Code %s." % (
                            employee.emp_code)))
                    time_off_data.append((0, 0, {
                        'request_date_from': leave['Leave/Start Date'],
                        'request_date_to': leave['Leave/End Date'],
                        'request_unit_half': True if leave['Leave/Half Day'] == 0.5 else False,
                        'employee_id': employee.id,
                        'holiday_status_id': leave_type.id,
                    }))
                payslip = self.env['hr.payslip'].create({
                    'name': payslip_name,
                    'select_month': self.select_month,
                    'select_year': self.select_year.id,
                    'date_from': self.date_from,
                    'date_to': self.date_to,
                    'employee_id': employee.id,
                    'emp_code': employee.emp_code,
                    'employee_manual_present_days': emp_data['Present Days'],
                    'extra_worked_days': emp_data['Extra Worked Days'],
                    'other_allowance': emp_data['Other Allowance'],
                    'attendance_bonus': emp_data['Attendance Bonus'],
                    'food_allowance': emp_data['Food Allowance'],
                    'incentive': emp_data['Incentive'],
                    'leave_encashment': emp_data['Leave Encashment'],
                    'overtime_hours': emp_data['Overtime Hours'],
                    'night_shift_allowance': emp_data['Night shift Allowance per day'],
                    'total_days_worked': emp_data['Total Night shift days worked'],
                    'production_incentive': emp_data['Production Incentive'],
                    'other_loan': emp_data['Other Loan'],
                    'food_deduction': emp_data['Food Deduction'],
                    'other_deduction': emp_data['Other Deduction'],
                    'room_rent_deduction': emp_data['Room Rent Deduction'],
                    'shoe_and_uniform_deduction': emp_data['Shoe & Uniform Deduction'],
                    'income_tax': emp_data['Income Tax'],
                    'actual_professional_tax': emp_data['Actual Professional Tax'],
                    'tds': emp_data['TDS'],
                    'time_off_ids': time_off_data,
                })
                payslip.employee_days_onchange()
                payslip.compute_sheet()
                payslip.compute_sheet()
                payslip.compute_sheet()
                payslip.compute_sheet()
                payslip.compute_sheet()
                self.write({
                    'payslip_ids': [(4, payslip.id)],
                    'status': 'generate'
                })

    # OLD CODE WITHOUT TIME OFF
    # def get_excel_data(self):
    #     self.ensure_one()
    #     if not self.excel_sheet:
    #         raise UserError(_("Please upload an Excel sheet first."))
    #     else:
    #         excel_data = BytesIO(base64.b64decode(self.excel_sheet))
    #         try:
    #             wb = load_workbook(excel_data)
    #             for ws in wb.worksheets:
    #                 for row in ws.iter_rows():
    #                     for cell in row:
    #                         cell.style = 'Normal'
    #                         if isinstance(cell.value, str) and cell.value.endswith('px'):
    #                             cell.value = cell.value.replace('px', '')
    #
    #             cleaned_data = BytesIO()
    #             wb.save(cleaned_data)
    #             cleaned_data.seek(0)
    #             df = pd.read_excel(cleaned_data)
    #         except Exception as e:
    #             raise UserError(_("Error reading Excel file: %s") % str(e))
    #         df.fillna(0, inplace=True)
    #         data_dict = df.to_dict(orient='records')
    #         table_header = ''
    #         table_datas = ''
    #         keys = df.columns.tolist()
    #         for key in keys:
    #             table_header += "<th style='border:1px solid black !important'>%s</th>" % key
    #         for row in data_dict:
    #             table_datas += "<tr>"
    #             for key in keys:
    #                 table_datas += "<td style='border:1px solid black !important'>{0}</td>".format(row[key])
    #             table_datas += "</tr>"
    #         html_table = """
    #                             <div style="overflow:auto;">
    #                                 <table class="table text-center table-border table-sm" style="width:max-content;">
    #                                     <thead>
    #                                         <tr style='border:1px solid black !important;background: lightblue;'>
    #                                             {header}
    #                                         </tr>
    #                                     </thead>
    #                                     <tbody>
    #                                         {data}
    #                                     </tbody>
    #                                 </table>
    #                             </div>
    #                             """.format(header=table_header, data=table_datas)
    #         self.data_html = html_table
    #         self.is_executed = True
    #         return data_dict

    # def generate_payslip(self):
    #     val = self.get_excel_data()
    #     for i in val:
    #         emp_code = i['Employee Code']
    #         employees = self.env['hr.employee'].search([('emp_code', '=', emp_code)])
    #         for employee in employees:
    #             payslip_name = self.env['ir.sequence'].next_by_code('salary.slip') or '/'
    #             payslip = self.env['hr.payslip'].create({
    #                 'name': payslip_name,
    #                 'select_month': self.select_month,
    #                 'select_year': self.select_year.id,
    #                 'date_from': self.date_from,
    #                 'date_to': self.date_to,
    #                 'employee_id': employee.id,
    #                 'emp_code': employee.emp_code,
    #                 'employee_manual_present_days': i['Present Days'],
    #                 'other_allowance': i['Other Allowance'],
    #                 'food_allowance': i['Food Allowance'],
    #                 'incentive': i['Incentive'],
    #                 'leave_encashment': i['Leave Encashment'],
    #                 'overtime_hours': i['Overtime Hours'],
    #                 'night_shift_allowance': i['Night shift Allowance per day'],
    #                 'total_days_worked': i['Total Night shift days worked'],
    #                 'production_incentive': i['Production Incentive'],
    #                 'other_loan': i['Other Loan'],
    #                 'food_deduction': i['Food Deduction'],
    #                 'other_deduction': i['Other Deduction'],
    #                 'room_rent_deduction': i['Room Rent Deduction'],
    #                 'shoe_and_uniform_deduction': i['Shoe & Uniform Deduction'],
    #                 'income_tax': i['Income Tax'],
    #                 'actual_professional_tax': i['Actual Professional Tax'],
    #             })
    #             payslip.employee_days_onchange()
    #             payslip.compute_sheet()
    #             payslip.compute_sheet()
    #             payslip.compute_sheet()
    #             payslip.compute_sheet()
    #             payslip.compute_sheet()
    #             self.write({
    #                 'payslip_ids': [(4, payslip.id)],
    #                 'status': 'generate'
    #             })
